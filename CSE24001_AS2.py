'''
LAB 02: BL.SC.U4CSE24001
19/7/2026
I used chatgpt to understand questions such as A5 and A6 and I've used openpyxl library to read/write into excel files. I couldve used pandas but i went with this because i started with it initially.
'''
#A1
import statistics as stat
from numpy import linalg
import openpyxl as xl
import numpy as np
from numpy.linalg import matrix_rank
import math
import matplotlib.pyplot as plt
import pandas as pd
import seaborn as sns
X=[]
y=[]
def A1():
    global X
    global y
    wb=xl.load_workbook("Lab Session Data.xlsx")
    sheet=wb["Purchase data"]
    i=0
    #X=[]
    #y=[]
    '''
    for row in sheet.iter_rows(max_col=5,values_only=True):    
        if i==0:
            X.append(list(row))
        else:
            y.append(list(row))
        i+=1'''
    for row in sheet.iter_rows(min_row=2, values_only=True):
        X.append(list(row[1:4]))   
        y.append(row[4])          

    X = np.array(X)
    rank=matrix_rank(X)
    #print(matrix_rank(X))
    #print(X)
    #print("loll")
    #print(y)

    X_pinv = np.linalg.pinv(X)
    cost = X_pinv @ y
    #print(X_pinv)
    #print("Cost of Candy :", cost[0])
    #print("Cost of Mango :", cost[1])
    #print("Cost of Milk  :", cost[2])
    return rank,cost[0],cost[1],cost[2]

def A2():
    global X,y
    classifier={}
    for i in range(len(y)):
        if y[i]>200:
            classifier["C_"+str(i+1)]="RICH"
        else:
            classifier["C_"+str(i+1)]="POOR"
    return classifier
def mean(lst):
    s=sum(lst)
    return s/len(lst)

def var(lst):
    res=0
    mn=mean(lst)
    for i in lst:
        res+= (i-mn)**2
    res=res/len(lst)
    return res


    
def A3():
    wb=xl.load_workbook("Lab Session Data.xlsx")
    sheet=wb["IRCTC Stock Price"]
    prices=[]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[4] is not None:
            prices.append(row[4])          
    #print(mean(prices))
    #print(var(prices))
    #print(np.mean(prices))
    #print(np.var(prices))
    Wprices=[]
    Aprices=[]
    chg=[]
    Wedcount=0
    wedprofc=0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if (row[4] is not None) and (row[2]=="Wed"):
            Wprices.append(row[3]) 
            Wedcount+=1
            if row[8]>0:
                wedprofc+=1
        if (row[4] is not None) and (row[1]=="Apr"):
            Aprices.append(row[3])     
        if row[8] is not None:
            chg.append(row[8])
    total=len(chg)
    chg=list(filter(lambda x: x<0,chg))
    ploss=len(chg)/total
    Wed_profits=wedprofc/total
    prof_onwed=wedprofc/Wedcount
    print(Wed_profits)

    #print(mean(Wprices))
    #print(var(Wprices))

    #return mean(prices),var(prices),np.mean(prices),np.var(prices),mean(Wprices),var(Wprices),mean(Aprices),ploss,Wed_profits,prof_onwed

def A3plot():
    wb=xl.load_workbook("Lab Session Data.xlsx")
    sheet=wb["IRCTC Stock Price"]
    days = []
    chg = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[8] is not None:
            days.append(row[2])   # Day
            chg.append(row[8])    # Chg%

    plt.scatter(days, chg)
    plt.title("Chg% vs Day of Week")
    plt.xlabel("Day")
    plt.ylabel("Chg%")
    plt.show()
def A4():
    wb=xl.load_workbook("Lab Session Data.xlsx")
    sheet=wb["thyroid0387_UCI"]
    numcount=0
    miss=0
    for col in sheet.iter_cols(min_row=0,max_row=9173,max_col=31, values_only=True):
        for i in col [1:9171]:
            if isinstance(i,int):
                numcount+=1
            if i is None or i=="?":
                miss+=1
    print("Total nums",numcount)
    print("miss",miss)

def A5():
    wb = xl.load_workbook("Lab Session Data.xlsx")
    sheet = wb["thyroid0387_UCI"]
    v1=[]  # this is for the first rec
    v2=[] # this is for the 2nd rec
    for i in sheet[2]:
        if isinstance(i.value,str):
            v1.append(i.value.lower())
        else:
            v1.append(i.value)
    for i in sheet[3]:
        if isinstance(i.value,str):
            v2.append(i.value.lower())    
        else:
            v2.append(i.value)
    print(v1)
    print(v2)
    f11=0
    f01=0
    f10=0
    f00=0
    
    #v1=[i.lower() for i in v1]
    for a, b in zip(v1, v2):
        if a in ("t", "f") and b in ("t", "f"):
            if a == "t" and b == "t":
                f11 += 1
            elif a == "t" and b == "f":
                f10 += 1
            elif a == "f" and b == "t":
                f01 += 1
            else:
                f00 += 1

    jc = f11 / (f11 + f10 + f01)
    smc = (f11 + f00) / (f11 + f10 + f01 + f00)
    return jc, smc

def A6():
    wb = xl.load_workbook("Lab Session Data.xlsx")
    sheet = wb["thyroid0387_UCI"]

    v1 = []
    v2 = []

    for i in sheet[2]:
        x = i.value

        if x == "t":    # t-> 1 , f->0, m->1 , f->0 and eeverything else is 0
            v1.append(1)
        elif x == "f":
            v1.append(0)
        elif x == "M":
            v1.append(1)
        elif x == "F":
            v1.append(0)
        elif x == "?":
            v1.append(0)
        elif x == "other":
            v1.append(0)
        elif x == "NO CONDITION":
            v1.append(0)
        else:
            v1.append(x)

    for cell in sheet[3]:
        x = cell.value

        if x == "t":
            v2.append(1)
        elif x == "f":
            v2.append(0)
        elif x == "M":
            v2.append(1)
        elif x == "F":
            v2.append(0)
        elif x == "?":
            v2.append(0)
        elif x == "other":
            v2.append(0)
        elif x == "NO CONDITION":
            v2.append(0)
        else:
            v2.append(x)

    v1 = np.array(v1, dtype=float)
    v2 = np.array(v2, dtype=float)

    cosine = np.dot(v1, v2) / (np.linalg.norm(v1) * np.linalg.norm(v2))

    print(cosine)

def encode(x):
    if x == "t":
        return 1
    elif x == "f":
        return 0
    elif x == "M":
        return 1
    elif x == "F":
        return 0
    elif isinstance(x, str):
        return 0
    else:
        return x


def JC(v1, v2):
    f11 = 0
    f10 = 0
    f01 = 0

    for a, b in zip(v1, v2):
        if a in [0, 1] and b in [0, 1]:
            if a == 1 and b == 1:
                f11 += 1
            elif a == 1 and b == 0:
                f10 += 1
            elif a == 0 and b == 1:
                f01 += 1

    if (f11 + f10 + f01) == 0:
        return 0

    return f11 / (f11 + f10 + f01)


def SMC(v1, v2):
    f11 = 0
    f10 = 0
    f01 = 0
    f00 = 0

    for a, b in zip(v1, v2):
        if a in [0, 1] and b in [0, 1]:
            if a == 1 and b == 1:
                f11 += 1
            elif a == 1 and b == 0:
                f10 += 1
            elif a == 0 and b == 1:
                f01 += 1
            else:
                f00 += 1

    total = f11 + f10 + f01 + f00

    if total == 0:
        return 0

    return (f11 + f00) / total


def COS(v1, v2):
    v1 = np.array(v1, dtype=float)
    v2 = np.array(v2, dtype=float)

    if np.linalg.norm(v1) == 0 or np.linalg.norm(v2) == 0:
        return 0

    return np.dot(v1, v2) / (np.linalg.norm(v1) * np.linalg.norm(v2))



def A7():
    wb = xl.load_workbook("Lab Session Data.xlsx")
    sheet = wb["thyroid0387_UCI"]

    observations = []

    for row in sheet.iter_rows(min_row=2, max_row=21):
        temp = []

        for cell in row:
            temp.append(encode(cell.value))

        observations.append(temp)

    jc_matrix = []
    smc_matrix = []
    cos_matrix = []

    for i in range(20):

        jc_row = []
        smc_row = []
        cos_row = []

        for j in range(20):

            jc_row.append(JC(observations[i], observations[j]))
            smc_row.append(SMC(observations[i], observations[j]))
            cos_row.append(COS(observations[i], observations[j]))

        jc_matrix.append(jc_row)
        smc_matrix.append(smc_row)
        cos_matrix.append(cos_row)

    return jc_matrix, smc_matrix, cos_matrix

def A8():
    wb = xl.load_workbook("Lab Session Data.xlsx")
    sheet = wb["thyroid0387_UCI"]

    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(list(row))

    cols = len(data[0])

    numeric = [1,18,20,22,24,26,28]  # these columns contain numeric data

    for c in range(cols):

        values = []

        for r in range(len(data)):
            if data[r][c] != "?":
                values.append(data[r][c])

        if c in numeric:

            values = [float(i) for i in values]

            fill = mean(values)

        else:

            fill = stat.mode(values)

        for r in range(len(data)):
            if data[r][c] == "?":
                data[r][c] = fill

    return data
def A9():

    data = A8()

    numeric = [1,18,20,22,24,26,28]

    for c in numeric:

        values = [float(row[c]) for row in data]

        minimum = min(values)
        maximum = max(values)

        for r in range(len(data)):

            if maximum == minimum:
                data[r][c] = 0

            else:
                data[r][c] = (float(data[r][c])-minimum)/(maximum-minimum)

    return data

# THE  MAIN FUNCTION 
print("========== A1 ==========")
R, candy, mango, milk = A1()
print("Rank of Feature Matrix :", R)
print("Cost of Candy          :", candy)
print("Cost of Mango          :", mango)
print("Cost of Milk           :", milk)

print("\n========== A2 ==========")
classified = A2()
for k, v in classified.items():
    print(k, ":", v)

print("\n========== A3 ==========")
A3()
A3plot()

print("\n========== A4 ==========")
A4()

print("\n========== A5 ==========")
jc, smc = A5()
print("Jaccard Coefficient         :", jc)
print("Simple Matching Coefficient :", smc)

print("\n========== A6 ==========")
A6()

print("\n========== A7 ==========")
jc_matrix, smc_matrix, cos_matrix = A7()

plt.figure(figsize=(8,6))
sns.heatmap(jc_matrix, annot=True)
plt.title("Jaccard Coefficient Heatmap")
plt.show()

plt.figure(figsize=(8,6))
sns.heatmap(smc_matrix, annot=True)
plt.title("Simple Matching Coefficient Heatmap")
plt.show()

plt.figure(figsize=(8,6))
sns.heatmap(cos_matrix, annot=True)
plt.title("Cosine Similarity Heatmap")
plt.show()

print("\n========== A8 ==========")
imputed = A8()
print("First 5 rows after Imputation:")
for row in imputed[:5]:
    print(row)

print("\n========== A9 ==========")
normalized = A9()
print("First 5 rows after Normalization:")
for row in normalized[:5]:
    print(row)